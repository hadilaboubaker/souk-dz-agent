#!/usr/bin/env python3
"""Standalone helper to sort a list of products by price and export to Excel.

Usage:
    python scripts/organiser_produits.py            # interactive prompts
    python scripts/organiser_produits.py --help     # see all options

The script accepts either:
- The bundled sample products (default), or
- A CSV file containing the columns: nom, prix, categorie, vendeur, lien

The output is an .xlsx workbook with one "Produits" sheet, sorted by ascending
price, with an extra "rang" column. A small per-category price summary is
printed to stdout.

This file intentionally has no dependency on the rest of `souk_dz` so it can
run on its own without configuring the full agent.
"""
from __future__ import annotations

import argparse
import csv
import statistics
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PRODUITS_EXEMPLE: list[dict[str, object]] = [
    {"nom": "iPhone 13", "prix": 450, "categorie": "Smartphone", "vendeur": "Jean", "lien": "https://fb.com/1"},
    {"nom": "Samsung Galaxy S21", "prix": 380, "categorie": "Smartphone", "vendeur": "Marie", "lien": "https://fb.com/2"},
    {"nom": "MacBook Pro 2019", "prix": 1200, "categorie": "Ordinateur", "vendeur": "Pierre", "lien": "https://fb.com/3"},
    {"nom": "Dell XPS 15", "prix": 850, "categorie": "Ordinateur", "vendeur": "Sophie", "lien": "https://fb.com/4"},
    {"nom": "PlayStation 5", "prix": 420, "categorie": "Console", "vendeur": "Luc", "lien": "https://fb.com/5"},
    {"nom": "Xbox Series X", "prix": 390, "categorie": "Console", "vendeur": "Emma", "lien": "https://fb.com/6"},
    {"nom": "iPad Air 4", "prix": 320, "categorie": "Tablette", "vendeur": "Thomas", "lien": "https://fb.com/7"},
    {"nom": "Samsung Tab S7", "prix": 350, "categorie": "Tablette", "vendeur": "Chloe", "lien": "https://fb.com/8"},
]

COLONNES = ["rang", "nom", "categorie", "prix", "vendeur", "lien"]

HEADER_FILL = PatternFill(start_color="0E4A8A", end_color="0E4A8A", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def charger_depuis_csv(fichier_csv: str | Path) -> list[dict[str, object]]:
    """Load products from a CSV file. Returns [] on error so the caller can fall back."""
    chemin = Path(fichier_csv)
    if not chemin.exists():
        print(f"✗ Fichier introuvable : {chemin}")
        return []
    try:
        with chemin.open(newline="", encoding="utf-8") as fh:
            reader = csv.DictReader(fh)
            produits: list[dict[str, object]] = []
            for row in reader:
                row = {k: (v if v != "" else None) for k, v in row.items()}
                if row.get("prix") is not None:
                    try:
                        row["prix"] = float(row["prix"])
                    except (TypeError, ValueError):
                        row["prix"] = None
                produits.append(row)
        print(f"✓ {len(produits)} produits chargés depuis {chemin}")
        return produits
    except Exception as exc:
        print(f"✗ Erreur lors du chargement : {exc}")
        return []


def _trier_et_indexer(produits: list[dict[str, object]]) -> list[dict[str, object]]:
    def _prix(row: dict[str, object]) -> float:
        valeur = row.get("prix")
        if isinstance(valeur, (int, float)):
            return float(valeur)
        return float("inf")

    tries = sorted(produits, key=_prix)
    for rang, row in enumerate(tries, start=1):
        row["rang"] = rang
    return tries


def _autosize(ws) -> None:
    for col_idx, col in enumerate(ws.columns, start=1):
        width = 12
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            width = min(50, max(width, len(value) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def organiser_produits_par_prix(
    produits: list[dict[str, object]],
    nom_fichier: str | Path = "produits_tries.xlsx",
) -> Path:
    """Sort the products by ascending price and write them to an Excel workbook.

    Returns the path to the file that was written.
    """
    if not produits:
        raise ValueError("Aucun produit à organiser.")

    tries = _trier_et_indexer([dict(p) for p in produits])

    wb = Workbook()
    ws = wb.active
    ws.title = "Produits"

    ws.append(COLONNES)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in tries:
        ws.append([row.get(col, "") for col in COLONNES])
    _autosize(ws)

    chemin = Path(nom_fichier)
    chemin.parent.mkdir(parents=True, exist_ok=True)
    wb.save(chemin)

    print(f"✓ Fichier Excel créé : {chemin}")
    print(f"✓ {len(tries)} produits organisés par prix croissant")
    _afficher_resume(tries)

    return chemin


def _afficher_resume(produits: list[dict[str, object]]) -> None:
    print("\n📊 Résumé par catégorie :")
    par_categorie: dict[str, list[float]] = defaultdict(list)
    for row in produits:
        prix = row.get("prix")
        if isinstance(prix, (int, float)):
            par_categorie[str(row.get("categorie") or "—")].append(float(prix))

    if not par_categorie:
        print("(aucun prix numérique disponible)")
        return

    largeur = max(len("categorie"), max(len(cat) for cat in par_categorie))
    entete = f"{'categorie'.ljust(largeur)}  {'min':>10}  {'max':>10}  {'moyenne':>10}  {'count':>6}"
    print(entete)
    print("-" * len(entete))
    for categorie in sorted(par_categorie):
        prix_list = par_categorie[categorie]
        print(
            f"{categorie.ljust(largeur)}  "
            f"{min(prix_list):>10.2f}  "
            f"{max(prix_list):>10.2f}  "
            f"{statistics.fmean(prix_list):>10.2f}  "
            f"{len(prix_list):>6d}"
        )


def _parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Trie une liste de produits par prix croissant et exporte vers Excel.",
    )
    parser.add_argument(
        "--csv",
        type=Path,
        help="Charge les produits depuis un fichier CSV (colonnes attendues : nom, prix, categorie, vendeur, lien).",
    )
    parser.add_argument(
        "--output",
        "-o",
        type=Path,
        default=Path("produits_tries.xlsx"),
        help="Nom du fichier Excel de sortie (défaut: produits_tries.xlsx).",
    )
    parser.add_argument(
        "--non-interactive",
        action="store_true",
        help="N'affiche aucun prompt ; utilise uniquement les arguments fournis.",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = _parse_args(argv)

    print("=== Organisateur de Produits par Prix ===\n")

    donnees: list[dict[str, object]]
    if args.csv:
        donnees = charger_depuis_csv(args.csv) or PRODUITS_EXEMPLE
    elif args.non_interactive:
        donnees = PRODUITS_EXEMPLE
    else:
        print("1. Utiliser les données d'exemple")
        print("2. Charger depuis un fichier CSV")
        choix = input("Choisissez une option (1 ou 2) : ").strip()
        if choix == "2":
            fichier = input("Entrez le nom du fichier CSV : ").strip()
            donnees = charger_depuis_csv(fichier) if fichier else []
            if not donnees:
                print("Utilisation des données d'exemple à la place.")
                donnees = PRODUITS_EXEMPLE
        else:
            donnees = PRODUITS_EXEMPLE

    nom_sortie: Path = args.output
    if not args.csv and not args.non_interactive:
        saisi = input(f"Nom du fichier de sortie (défaut: {nom_sortie}) : ").strip()
        if saisi:
            nom_sortie = Path(saisi)

    organiser_produits_par_prix(donnees, nom_sortie)
    return 0


if __name__ == "__main__":
    sys.exit(main())
