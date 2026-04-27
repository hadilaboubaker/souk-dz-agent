"""Build an Excel attachment with all listings + the top opportunities."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from souk_dz.models import NormalizedListing, Opportunity

HEADER_FILL = PatternFill(start_color="0E4A8A", end_color="0E4A8A", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
OPP_FILL = PatternFill(start_color="FFF3C4", end_color="FFF3C4", fill_type="solid")


def _autosize(ws):
    for col_idx, col in enumerate(ws.columns, start=1):
        width = 12
        for cell in col:
            try:
                value = "" if cell.value is None else str(cell.value)
                width = min(60, max(width, len(value) + 2))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_excel(
    out_path: Path,
    all_items: list[NormalizedListing],
    opportunities: list[Opportunity],
) -> Path:
    wb = Workbook()

    # ---- Sheet 1: Opportunities ----
    ws = wb.active
    ws.title = "الفرص"
    headers = [
        "#",
        "اسم المنتج",
        "Catégorie",
        "السعر (DZD)",
        "متوسط السوق (DZD)",
        "الخصم %",
        "المصدر",
        "الولاية",
        "التواصل",
        "Lien",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for idx, opp in enumerate(opportunities, start=1):
        listing = opp.listing.listing
        ws.append(
            [
                idx,
                opp.listing.canonical_name,
                opp.listing.category,
                listing.price_dzd or 0,
                round(opp.median_price_dzd),
                round(opp.discount_percent, 1),
                listing.source_label,
                listing.wilaya or "",
                listing.contact or "",
                str(listing.url) if listing.url else "",
            ]
        )
        for cell in ws[idx + 1]:
            cell.fill = OPP_FILL
    _autosize(ws)

    # ---- Sheet 2: All listings ----
    ws2 = wb.create_sheet("كل الإعلانات")
    headers2 = [
        "#",
        "اسم المنتج",
        "Catégorie",
        "Brand",
        "السعر (DZD)",
        "السعر المنشور",
        "المصدر",
        "الولاية",
        "Cluster key",
        "Lien",
    ]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for idx, item in enumerate(all_items, start=1):
        listing = item.listing
        ws2.append(
            [
                idx,
                item.canonical_name,
                item.category,
                item.brand or "",
                listing.price_dzd or "",
                listing.price_raw or "",
                listing.source_label,
                listing.wilaya or "",
                item.cluster_key,
                str(listing.url) if listing.url else "",
            ]
        )
    _autosize(ws2)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
